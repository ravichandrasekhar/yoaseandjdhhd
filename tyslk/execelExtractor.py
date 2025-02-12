import io
from pydantic import BaseModel
from openpyxl import load_workbook
class ExecelExtractor:
    
     class InputSchema(BaseModel):
         InputSchema: str
     def process_xlsx(xlsx_content):
        """Extract and chunk text from an XLSX file."""
        wb = load_workbook(io.BytesIO(xlsx_content), data_only=True)
        chunks = ''
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    chunks=chunks+cell.value
                    
        return chunks