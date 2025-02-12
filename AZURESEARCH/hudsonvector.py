from openai import AzureOpenAI
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.models import VectorizedQuery
import openpyxl
client = AzureOpenAI(
    azure_deployment="text-embedding-ada-002",
    api_version="2023-05-15",
    azure_endpoint="",
    api_key=""
)

# Define your Azure Cognitive Search endpoint and key
service_endpoint = ""
api_key = ""

# Initialize a SearchClient
credential = AzureKeyCredential(api_key)
search_client = SearchClient(endpoint=service_endpoint, index_name="hudsonsharepointlist", credential=credential)

# Pure Vector Search

query = "what are the responsibilites of Hudson Advisors?"
embedding = client.embeddings.create(input=query, model="text-embedding-ada-002").data[0].embedding
vector_query = VectorizedQuery(vector=embedding, k_nearest_neighbors=3, fields="ContentEmbeddings")
 
results = search_client.search(  
    search_text=query,  
    vector_queries=[vector_query],
    select=["id", "question", "answer"],
)  

# Print the top 3 results
for result in results: 
    print(f"Score: {result['@search.score']}")
    print(f"id:{result['id']}")   
    print(f"Question: {result['question']}")  
    print(f"Answer: {result['answer']}")
# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Search Results"

# Write headers
headers = ["Score", "id", "question", "answer"]
for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)

# Write search results
row = 2
for result in results: 
    sheet.cell(row=row, column=1, value=result['@search.score'])
    sheet.cell(row=row, column=2, value=result['id'])
    sheet.cell(row=row, column=3, value=result['question'])
    sheet.cell(row=row, column=4, value=result['answer'])
    row += 1

# Save the workbook
workbook.save("search_results.xlsx")
print("Search results saved to search_results.xlsx")
print(workbook.save("search_results.xlsx"))