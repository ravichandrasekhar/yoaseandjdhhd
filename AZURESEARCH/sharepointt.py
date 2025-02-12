import requests
import re
import os
import io
import pdfplumber
import openpyxl
import csv
from docx import Document
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import SearchIndex, SimpleField, SearchableField, SearchField
from azure.search.documents.indexes.models import HnswAlgorithmConfiguration, VectorSearch, VectorSearchAlgorithmConfiguration, VectorSearchProfile
from openai import AzureOpenAI
from azure.identity import DefaultAzureCredential, get_bearer_token_provider

# Constants
SHAREPOINT_SITE_ID = ""
ACCESS_TOKEN = ""
list_id = ""
group_id = ""
folder_id = ""
SHAREPOINT_DRIVE_ID = ""

# Azure search
SEARCH_ENDPOINT = ""
SEARCH_ADMIN_KEY = ""
SEARCH_INDEX_NAME = "sharepoint"

# Azure OpenAI
azure_openai_endpoint = ""
azure_openai_key = ""
azure_openai_embedding_deployment = "text-embedding-ada-002"
azure_openai_version = "2023-05-15"

def regularexpression(field_name):
    """Sanitize the field name to comply with Azure Cognitive Search naming rules."""
    # Remove invalid characters
    expression = re.sub(r'[^a-zA-Z0-9_]', '_', field_name)
    # Ensure the name starts with a letter
    if not re.match(r'^[a-zA-Z]', expression):
        expression = 'f_' + expression
    return expression

def create_search_index(fields):
    try:
        index_client = SearchIndexClient(endpoint=SEARCH_ENDPOINT, credential=AzureKeyCredential(SEARCH_ADMIN_KEY))
        index_fields = [
            SimpleField(name="id", type="Edm.String", key=True),
            SimpleField(name="webUrl", type="Edm.String"),
            SearchField(name="contentEmbeddings", type="Collection(Edm.Single)", vector_search_dimensions=1536, vector_search_profile_name="myHnswProfile"),
            SimpleField(name="Permissions", type="Collection(Edm.String)"),
            SimpleField(name="accessList", type="Collection(Edm.String)")
        ]

        for field in fields:
            regularxname = regularexpression(field)
            index_fields.append(SimpleField(name=regularxname, type="Edm.String"))
        
        vector_search = VectorSearch(
            algorithms=[HnswAlgorithmConfiguration(name="myHnsw")],
            profiles=[VectorSearchProfile(name="myHnswProfile", algorithm_configuration_name="myHnsw")]
        )
        
        index = SearchIndex(name=SEARCH_INDEX_NAME, fields=index_fields, vector_search=vector_search)
        result = index_client.create_index(index)
        print("Index created:", result)
        
    except Exception as e:
        print(f"Error creating search index: {str(e)}")

def get_sharepoint_list_data(list_id):
    try:
        schema_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/lists/{list_id}/columns"
        headers = {"Authorization": "Bearer " + ACCESS_TOKEN, "Accept": "application/json"}
        response = requests.get(schema_url, headers=headers)
        
        if response.status_code == 200:
            schema_data = response.json()
            fields_to_expand_and_index = [column["name"] for column in schema_data["value"]]
            select_fields = ','.join(fields_to_expand_and_index)
            url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/lists/{list_id}/items?expand=fields(select={select_fields})"
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                return data['value'], fields_to_expand_and_index
            else:
                print(f"Failed to retrieve data from SharePoint list. Status code: {response.status_code}")
                return [], []
        else:
            print(f"Failed to retrieve schema from SharePoint list. Status code: {response.status_code}")
            return [], []
    except Exception as e:
        print(f"Error fetching data from SharePoint list: {str(e)}")
        return [], []

def get_item_permissions(item_id):
    try:
        permissions_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/lists/{list_id}/items/{item_id}/driveItem/permissions"
        headers = {"Authorization": "Bearer " + ACCESS_TOKEN, "Accept": "application/json"}
        response = requests.get(permissions_url, headers=headers)
        permissions = []
        
        if response.status_code == 200:
            permissions_data = response.json().get("value", [])
            for permission_data in permissions_data:
                granted_to = permission_data.get("grantedTo", None)
                if granted_to:
                    if "user" in granted_to:
                        user_or_group = granted_to["user"].get("displayName", "N/A")
                        permissions.append(user_or_group)
                    elif "group" in granted_to:
                        user_or_group = granted_to["group"].get("displayName", "N/A")
                        permissions.append(user_or_group)
        return permissions
    except Exception as e:
        print(f"Error fetching permissions for item {item_id}: {str(e)}")
        return []

def get_group_members(group_id):
    try:
        url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/transitiveMembers/"
        group_members = []
        headers = {"Authorization": "Bearer " + ACCESS_TOKEN, "Accept": "application/json"}
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            members = response.json()["value"]
            for member in members:
                principal_name = member["userPrincipalName"]
                group_members.append(principal_name)
        return group_members
    except Exception as e:
        print(f"Error fetching group members for group {group_id}: {str(e)}")
        return []

def get_sharepoint_items(access_token, folder_id=None):
    if folder_id:
        url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{folder_id}/children"
    else:
        url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/root/children"

    headers = {"Authorization": "Bearer " + access_token, "Accept": "application/json"}
    response = requests.get(url, headers=headers)
    data = response.json()
    return data['value']

def download_sharepoint_file(file_id, access_token):
    content_url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{file_id}/content"
    headers = {"Authorization": "Bearer " + access_token}
    response = requests.get(content_url, headers=headers)
    
    if response.status_code == 200:
        return response.content
    else:
        print(f"Failed to download file with ID '{file_id}'")
        return None

def get_permissions(file_id, access_token):
    permissions_url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{file_id}/permissions"
    headers = {"Authorization": "Bearer " + access_token, "Accept": "application/json"}
    response = requests.get(permissions_url, headers=headers)
    permissions = []
    
    if response.status_code == 200:
        permissions_data = response.json().get("value", [])
        for permission_data in permissions_data:
            granted_to = permission_data.get("grantedTo", None)
            if granted_to:
                if "user" in granted_to:
                    user_or_group = granted_to["user"].get("displayName", "N/A")
                elif "group" in granted_to:
                    user_or_group = granted_to["group"].get("displayName", "N/A")
                if user_or_group != "N/A":  # Filter out "N/A" values
                    permissions.append(f" {user_or_group}")

    return permissions

def extract_text_from_pdf(file_content):
    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def extract_text_from_word(file_content):
    doc = Document(io.BytesIO(file_content))
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def extract_text_from_excel(file_content):
    file_str = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        file_str += str(cell.value) + " "
    except Exception as e:
        file_str = f"Error extracting text from Excel file: {str(e)}"
    return file_str

def extract_text_from_text(file_content):
    try:
        return file_content.decode('utf-8')
    except UnicodeDecodeError:
        return str(file_content, errors='ignore')

def extract_text_from_csv(file_content):
    file_str = ""
    try:
        content_str = file_content.decode('utf-8')
        reader = csv.reader(content_str.splitlines())
        for row in reader:
            for cell in row:
                file_str += cell + " "
    except Exception as e:
        file_str = f"Error extracting text from CSV file: {str(e)}"
    return file_str

def extract_text_from_other(file_content):
    try:
        file_str = file_content.decode("utf-8")
    except UnicodeDecodeError:
        file_str = "Unable to decode file content"
    return file_str

def generate_embeddings(content):
    try:
        openai_credential = DefaultAzureCredential()
        token_provider = get_bearer_token_provider(openai_credential, "https://cognitiveservices.azure.com/.default")
        client = AzureOpenAI(
            azure_deployment=azure_openai_embedding_deployment,
            azure_endpoint=azure_openai_endpoint,
            api_key=azure_openai_key,
            azure_ad_token_provider=token_provider if not azure_openai_key else None,
            api_version=azure_openai_version
        )

        response = client.embeddings.create(input=content, model="text-embedding-ada-002")
        embeddings = response.data[0].embedding
        return embeddings
    except Exception as e:
        print(f"Error generating embeddings for content: {e}")
        return []

def index_documents_to_azure_search(all_documents):
    try:
        client_search = SearchClient(endpoint=SEARCH_ENDPOINT, index_name=SEARCH_INDEX_NAME, credential=AzureKeyCredential(SEARCH_ADMIN_KEY))

        for document in all_documents:
            content = document.get('Content', '')
            embeddings = generate_embeddings(content)
            document['contentEmbeddings'] = embeddings

            try:
                client_search.upload_documents(documents=[document])
                print(f"Document '{document['id']}' indexed successfully.")
            except Exception as e:
                print(f"Failed to index document '{document['id']}': {str(e)}")
    except Exception as e:
        print(f"Failed to index documents to Azure Cognitive Search. Error: {str(e)}")

def index_folder_to_azure_search(folder_name, folder_id):
    folder_items = get_sharepoint_items(ACCESS_TOKEN, folder_id)
    all_documents = []
    for item in folder_items:
        if 'folder' in item:
            index_folder_to_azure_search(item['name'], item['id'])
        else:
            file_content = download_sharepoint_file(item['id'], ACCESS_TOKEN)
            permissions = get_permissions(item['id'], ACCESS_TOKEN)
            group_members = get_group_members(group_id, ACCESS_TOKEN)
            document = {
                "id": str(item['id']),
                "webUrl": item['webUrl'],
                "Permissions": permissions,
                "accessList": group_members,
                "ContentEmbeddings": [],
                "Content": extract_text_from_other(file_content)
            }
            all_documents.append(document)
    index_documents_to_azure_search(all_documents)

if __name__ == "__main__":
    try:
        # Index SharePoint List
        if list_id:
            sharepoint_data, fields = get_sharepoint_list_data(list_id)
            if sharepoint_data:
                create_search_index(fields)
                all_documents = []
                for item in sharepoint_data:
                    document = {
                        "id": str(item['id']),
                        "webUrl": item['webUrl'],
                        "Permissions": get_item_permissions(item['id']),
                        "accessList": get_group_members(group_id),
                        "ContentEmbeddings": [],
                    }
                    for field in fields:
                        rexgufield = regularexpression(field)
                        document[rexgufield] = str(item.get('fields', {}).get(field, ''))
                    all_documents.append(document)
                index_documents_to_azure_search(all_documents)
            else:
                print("No data retrieved from SharePoint list.")
        
        # Index SharePoint Drive
        sharepoint_drive_items = get_sharepoint_items(ACCESS_TOKEN)
        all_documents = []
        for item in sharepoint_drive_items:
            if 'folder' in item:
                index_folder_to_azure_search(item['name'], item['id'])
                
            else:
                file_content = download_sharepoint_file(item['id'], ACCESS_TOKEN)
                permissions = get_permissions(item['id'], ACCESS_TOKEN)
                group_members = get_group_members(group_id, ACCESS_TOKEN)
                document = {
                    "id": str(item['id']),
                    "webUrl": item['webUrl'],
                    "Permissions": permissions,
                    "accessList": group_members,
                    "ContentEmbeddings": [],
                    "Content": extract_text_from_other(file_content)
                }
                all_documents.append(document)
        index_documents_to_azure_search(all_documents)
    except Exception as e:
        print(f"Main process failed: {str(e)}")
