from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from typing import Optional
import requests
import uvicorn
import pdfplumber
from docx import Document
import io
import os
import csv
import openpyxl
from fastapi import HTTPException
from azure.core.exceptions import HttpResponseError
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import SearchIndex, SearchField, SimpleField, SearchableField, SearchFieldDataType
from azure.core.credentials import AzureKeyCredential

app = FastAPI()
search_service_endpoint = ""
admin_key = ""
index_name = "sharepoint-list-index"
group_id = "f24f4eeb-5714-436d-88ff-e5c251399c91"

# Define request models
class SharePointItem(BaseModel):
    ACCESS_TOKEN: Optional[str] = Field(None, description="Access token for SharePoint API")
    CLIENT_ID: Optional[str] = Field(None, description="Client ID for SharePoint API")
    CLIENT_SECRET: Optional[str] = Field(None, description="Client secret for SharePoint API")
    TENANT_ID: Optional[str] = Field(None, description="Tenant ID for SharePoint API")
    configuration: str = Field(..., description="Configuration type: 'list', 'drive', or 'pages'")
    SHAREPOINT_DRIVE_ID: Optional[str] = Field(None, description="ID of the SharePoint drive (optional)")
    list_id: Optional[str] = Field(None, description="ID of the SharePoint list (optional)")
    sharepoint_site_id: Optional[str] = Field(None, description="ID of the SharePoint site (optional)")
    folder_id: Optional[str] = Field(None, description="ID of the SharePoint folder (optional)")
def sanitize_field_name(field_name):
    # Function to sanitize field names according to Azure Search schema
    return field_name.replace(" ", "_").replace(".", "_")
def sanitize_field_name(field_name):
    if field_name.startswith('_'):
        return field_name[1:]
    return field_name    

def create_or_update_index(search_service_endpoint, admin_key, index_name, fields):
    client = SearchIndexClient(endpoint=search_service_endpoint, credential=AzureKeyCredential(admin_key))
    field_schemas = [
        SimpleField(name=sanitize_field_name(field), type=SearchFieldDataType.String, searchable=True, filterable=True)
        for field in fields
    ]
    field_schemas.append(SimpleField(name="id", type=SearchFieldDataType.String, key=True))
    
    field_schemas.append(SimpleField(name="accessList", type="Collection(Edm.String)"))
    field_schemas.append(SimpleField(name="permissions", type="Collection(Edm.String)"))
    
    index = SearchIndex(name=index_name, fields=field_schemas)
    
    client.create_or_update_index(index)
    print(f"Search Index successfully created {index_name}")

def get_item_permissions(item_id, access_token, sharepoint_site_id, list_id):
    try:
        permissions_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/items/{item_id}/driveItem/permissions"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(permissions_url, headers=headers)
        permissions = []
        if response.status_code == 200:
            permissions_data = response.json().get("value", [])
            for permission_data in permissions_data:
                granted_to = permission_data.get("grantedTo", None)
                if granted_to:
                    if "user" in granted_to:
                        user_or_group = granted_to["user"].get("displayName", "N/A")
                        permissions.append(user_or_group)
                    elif "group" in granted_to:
                        user_or_group = granted_to["group"].get("displayName", "N/A")
                        permissions.append(user_or_group)

        return permissions
    except Exception as e:
        print(f"Error fetching permissions for item {item_id}: {str(e)}")
        return []

def get_group_members(group_id, access_token):
    try:
        url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/transitiveMembers/"
        group_members = []
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            members = response.json()["value"]

            for member in members:
                principal_name = member["userPrincipalName"]
                group_members.append(principal_name)

        return group_members
    except Exception as e:
        print(f"Error fetching group members for group {group_id}: {str(e)}")
        return []

def upload_documents(search_service_endpoint, admin_key, index_name, documents):
   try:
        client = SearchClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(admin_key))
        
        # Convert numerical values to strings in document data
        for document in documents:
            for key, value in document.items():
                if isinstance(value, (int, float)):
                    document[key] = str(value)
        
        client.upload_documents(documents)
        print("Documents uploaded successfully.")
   except Exception as e:
        print(f"An error occurred while uploading documents: {str(e)}")

def get_access_token(client_id, client_secret, tenant_id):
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }

    response = requests.post(token_url, data=data)
    if response.status_code == 200:
        return response.json().get("access_token")
    else:
        print(f"Failed to retrieve access token: {response.text}")
        return None

def get_sharepoint_list_data(list_id, access_token, sharepoint_site_id):
    try:
        schema_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/columns"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(schema_url, headers=headers)
        if response.status_code == 200:
            schema_data = response.json()
            fields_to_expand_and_index = [column["name"] for column in schema_data["value"]]

            select_fields = ','.join(fields_to_expand_and_index)
            url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/items?expand=fields(select={select_fields})"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                if 'value' in data:
                    return data['value'], fields_to_expand_and_index
                else:
                    print("No items found in the SharePoint list.")
                    return [], []
            else:
                print(f"Failed to retrieve data from SharePoint list. Status code: {response.status_code}")
                return [], []
        else:
            print(f"Failed to retrieve schema from SharePoint list. Status code: {response.status_code}")
            return [], []
    except Exception as e:
        print(f"Error fetching data from SharePoint list: {str(e)}")
        return [], []

def get_all_sharepoint_lists(access_token, sharepoint_site_id):
    try:
        lists_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(lists_url, headers=headers)
        if response.status_code == 200:
            lists_data = response.json().get("value", [])
            all_lists_data = []
            for list_item in lists_data:
                list_id = list_item.get("id")
                list_data, fields = get_sharepoint_list_data(list_id, access_token, sharepoint_site_id)
                all_lists_data.append({"list_id": list_id, "list_data": list_data, "fields": fields})
            return all_lists_data
        else:
            print(f"Failed to retrieve lists from SharePoint site. Status code: {response.status_code}")
            return []
    except Exception as e:
        print(f"Error fetching lists from SharePoint site: {str(e)}")
        return []

def get_sharepoint_items(access_token, folder_id, SHAREPOINT_DRIVE_ID):
    try:
        if folder_id:
            url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{folder_id}/children"
        else:
            url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/root/children"

        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            return data['value']
        else:
            raise HTTPException(status_code=response.status_code, detail="Failed to retrieve items from SharePoint drive.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching items from SharePoint drive: {str(e)}")

def download_sharepoint_file(access_token, item_id, SHAREPOINT_DRIVE_ID):
    try:
        url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{item_id}/content"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(url, headers=headers, stream=True)
        if response.status_code == 200:
            file_content = response.content
            return file_content
        else:
            raise HTTPException(status_code=response.status_code, detail="Failed to download file from SharePoint drive.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading file from SharePoint drive: {str(e)}")

def extract_text_from_pdf(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
        return text

def extract_text_from_docx(docx_bytes):
    doc = Document(io.BytesIO(docx_bytes))
    text = ""
    for para in doc.paragraphs:
        text += para.text + "\n"
    return text

def extract_text_from_xlsx(xlsx_bytes):
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    text = ""
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text += str(cell) + "\t"
            text += "\n"
    return text

def extract_text_from_csv(csv_bytes):
    text = ""
    with io.StringIO(csv_bytes.decode('utf-8')) as f:
        reader = csv.reader(f)
        for row in reader:
            text += ",".join(row) + "\n"
    return text

def extract_text_from_file(file_name, file_bytes):
    if file_name.lower().endswith('.pdf'):
        return extract_text_from_pdf(file_bytes)
    elif file_name.lower().endswith('.docx'):
        return extract_text_from_docx(file_bytes)
    elif file_name.lower().endswith('.xlsx'):
        return extract_text_from_xlsx(file_bytes)
    elif file_name.lower().endswith('.csv'):
        return extract_text_from_csv(file_bytes)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

def index_sharepoint_files(items, access_token, SHAREPOINT_DRIVE_ID):
    documents = []
    for item in items:
        if "file" in item:
            file_name = item["name"]
            item_id = item["id"]
            file_bytes = download_sharepoint_file(access_token, item_id, SHAREPOINT_DRIVE_ID)
            file_content = extract_text_from_file(file_name, file_bytes)

            # permissions = get_item_permissions(item_id, access_token, sharepoint_site_id, list_id)
            # accessList = get_group_members(group_id, access_token)
            
            document = {
                "id": item_id,
                "file_name": file_name,
                "content": file_content,
                # "accessList": accessList,
                # "permissions": permissions
            }
            documents.append(document)
    return documents

@app.post("/process_sharepoint_item/")
async def process_sharepoint_item(item: SharePointItem):
    try:
        if item.ACCESS_TOKEN:
            access_token = item.ACCESS_TOKEN
        else:
            access_token = get_access_token(item.CLIENT_ID, item.CLIENT_SECRET, item.TENANT_ID)
        
        if not access_token:
            raise HTTPException(status_code=500, detail="Failed to retrieve access token.")
        
        if item.configuration == "list":
            try:
                all_lists_data = get_all_sharepoint_lists(access_token, item.sharepoint_site_id)
                if all_lists_data:
                    for list_data in all_lists_data:
                        fields = list_data["fields"]
                        list_items = list_data["list_data"]

                        documents = []
                        for list_item in list_items:
                            item_id = list_item["id"]
                            permissions = get_item_permissions(item_id, access_token, item.sharepoint_site_id, list_data["list_id"])
                            accessList = get_group_members(group_id, access_token)
                            document = {
                                "id": list_item["id"],
                                "accessList": accessList,
                                "permissions": permissions
                            }
                            # Add only allowed fields
                            for field, value in list_item["fields"].items():
                                if field not in ["_UIVersionString", "another_undesired_field"]:
                                    document[sanitize_field_name(field)] = value
                            documents.append(document)

                        create_or_update_index(search_service_endpoint, admin_key, index_name, fields)
                        upload_documents(search_service_endpoint, admin_key, index_name, documents)

            except Exception as e:
                print(status_code=500, detail=f"Failed to process list items: {str(e)}")

        elif item.configuration == "drive":
            try:
                items = get_sharepoint_items(access_token, item.folder_id, item.SHAREPOINT_DRIVE_ID)
                documents = index_sharepoint_files(items, access_token, item.SHAREPOINT_DRIVE_ID)
                upload_documents(search_service_endpoint, admin_key, index_name, documents)
            except HttpResponseError as e:
                raise HTTPException(status_code=500, detail=f"Failed to process drive items: {str(e)}")

        elif item.configuration == "pages":
            try:
                items = get_sharepoint_items(access_token, item.folder_id, item.SHAREPOINT_DRIVE_ID)
                documents = index_sharepoint_files(items, access_token, item.SHAREPOINT_DRIVE_ID)
                upload_documents(search_service_endpoint, admin_key, index_name, documents)
            except HttpResponseError as e:
                raise HTTPException(status_code=500, detail=f"Failed to process pages items: {str(e)}")

        return {"message": "SharePoint item processed successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
