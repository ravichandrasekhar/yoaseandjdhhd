from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from typing import Optional, Dict
import requests
import uvicorn
import pdfplumber
from docx import Document
import io
import os
import csv
import openpyxl
import base64
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import SearchIndex, SimpleField, SearchField,SearchableField, HnswAlgorithmConfiguration, VectorSearch, VectorSearchAlgorithmConfiguration, VectorSearchProfile,SearchFieldDataType
from azure.core.credentials import AzureKeyCredential
from openai import AzureOpenAI
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
import re
app = FastAPI()
search_service_endpoint = ""
admin_key = ""
index_name = "sharepoint"
group_id = ""
azure_openai_endpoint = ""
azure_openai_key = ""
azure_openai_embedding_deployment = "text-embedding-ada-002"
embedding_model_name = "text-embedding-ada-002"
azure_openai_version = "2023-05-15"
# Define request models
class SharePointItem(BaseModel):
    ACCESS_TOKEN: Optional[str] = Field(None, description="Access token for SharePoint API")
    CLIENT_ID: Optional[str] = Field(None, description="Client ID for SharePoint API")
    CLIENT_SECRET: Optional[str] = Field(None, description="Client secret for SharePoint API")
    TENANT_ID: Optional[str] = Field(None, description="Tenant ID for SharePoint API")
    configuration: str = Field(..., description="Configuration type: 'list', 'drive', or 'pages'")
    SHAREPOINT_DRIVE_ID: Optional[str] = Field(None, description="ID of the SharePoint drive (optional)")
    list_id: Optional[str] = Field(None, description="ID of the SharePoint list (optional)")
    sharepoint_site_id: Optional[str] = Field(None, description="ID of the SharePoint site (optional)")
    folder_id: Optional[str] = Field(None, description="ID of the SharePoint folder (optional)")
def sanitize_field_names(fields):
    return [sanitize_field_name(field) for field in fields]

def sanitize_field_name(field_name):
    if field_name.startswith('_'):
        return field_name[1:]
    return field_name
def regularexpression(field_name):
    """Sanitize the field name to comply with Azure Cognitive Search naming rules."""
    # Remove invalid characters
    expression = re.sub(r'[^a-zA-Z0-9_]', '_', field_name)
    # Ensure the name starts with a letter
    if not re.match(r'^[a-zA-Z]', expression):
        expression = 'f_' + expression
    return expression    
def get_sharepoint_pages(site_id, access_token):
    graph_endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_id}/pages"
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    response = requests.get(graph_endpoint, headers=headers)
    if response.status_code == 200:
        return response.json().get("value", [])
    else:
        print(f"Failed to retrieve SharePoint pages data: {response.text}")
        return []
def create_or_update_index(search_service_endpoint, admin_key, index_name, fields):
    client = SearchIndexClient(endpoint=search_service_endpoint, credential=AzureKeyCredential(admin_key))
    index_fields = [
            SimpleField(name="id", type="Edm.String", key=True),
            SimpleField(name="Permissions", type="Collection(Edm.String)"),
            SimpleField(name="accessList", type="Collection(Edm.String)"),
            SearchField(name="contentEmbeddings", type="Collection(Edm.Single)",vector_search_dimensions=1536, vector_search_profile_name="myHnswProfile")
        ]
    for field in fields:
            regularxname = regularexpression(field)
            index_fields.append(SearchableField(name=regularxname, type="Edm.String"))
        
    vector_search = VectorSearch(
            algorithms=[
                HnswAlgorithmConfiguration(
                    name="myHnsw"
                )
            ],
            profiles=[
                VectorSearchProfile(
                    name="myHnswProfile",
                    algorithm_configuration_name="myHnsw",
                )
            ]
        )

    index = SearchIndex(name=index_name, fields=index_fields,vector_search=vector_search)
    index_client = SearchIndexClient(search_service_endpoint, AzureKeyCredential(admin_key))
    try:
        # Check if the index exists
        existing_index = index_client.get_index(index_name)
        print(f"Index '{index_name}' already exists. Deleting the existing index...")
        index_client.delete_index(index_name)
        print(f"Index '{index_name}' has been deleted.")
    except Exception as e:
        print(f"Index '{index_name}' not found.")

    try:
        print(f"Creating index '{index_name}'...")
        index_client.create_index(index)
        print(f"Index '{index_name}' has been created successfully.")
    except Exception as e:
        print(f"Error creating index '{index_name}': {e}")
#sharepoint drivecode
def get_sharepoint_items(access_token, folder_id, SHAREPOINT_DRIVE_ID):
    try:
        if folder_id:
            url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{folder_id}/children"
        else:
            url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/root/children"

        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            return data['value']
        else:
            raise HTTPException(status_code=response.status_code, detail="Failed to retrieve items from SharePoint drive.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching items from SharePoint drive: {str(e)}")

def download_sharepoint_file(file_id, SHAREPOINT_DRIVE_ID, access_token):
    try:
        content_url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{file_id}/content"
        headers = {
            "Authorization": "Bearer " + access_token,
        }
        response = requests.get(content_url, headers=headers)
        if response.status_code == 200:
            return response.content
        else:
            raise HTTPException(status_code=response.status_code, detail=f"Failed to download file with ID '{file_id}'")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading file from SharePoint drive: {str(e)}")

def get_sharepoint_pages(site_id, access_token):
    graph_endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_id}/pages"
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    response = requests.get(graph_endpoint, headers=headers)
    if response.status_code == 200:
        print(response.json().get("value", []))
        return response.json().get("value", [])
        
    else:
        print(f"Failed to retrieve SharePoint pages data: {response.text}")
        return []
def index_sharepoint_pages(service_endpoint, index_name, api_key, site_id, access_token, group_id):
    try:
        openai_credential = DefaultAzureCredential()
        token_provider = get_bearer_token_provider(openai_credential, "https://cognitiveservices.azure.com/.default")
        client = AzureOpenAI(
            azure_deployment=azure_openai_embedding_deployment,
            azure_endpoint=azure_openai_endpoint,
            api_key=azure_openai_key,
            azure_ad_token_provider=token_provider if not azure_openai_key else None,
            api_version=azure_openai_version
        )

        pages = get_sharepoint_pages(site_id, access_token)
        group_members = get_group_members(group_id, access_token)

        documents = []
        for page in pages:
            document = {
                "id": str(page["id"]),
                "title": page.get("title", ""),
                "description": page.get("description", ""),
                "webUrl":page.get("webUrl",""),
                "groupMembers": [member["name"] for member in group_members],
                "ContentEmbeddings":[]
            }
            content = document.get('description', '')  # Get the content from the description
            embeddings = []
            content_str = "\n".join(content.split('\n'))  # Concatenate content lines into a single string
            try:
                # Generate embeddings for content
                response = client.embeddings.create(input=content_str, model=embedding_model_name)
                embeddings = response.data[0].embedding  # Access embeddings directly, assuming response is in the correct format
                embeddings_str = [str(embedding) for embedding in embeddings]
                document['ContentEmbeddings'] = embeddings_str
            except Exception as e:
                print(f"Error generating embeddings for content '{content_str}': {e}")

            documents.append(document)
        
        client = SearchClient(endpoint=service_endpoint, index_name=index_name, credential=AzureKeyCredential(api_key))
        result = client.upload_documents(documents=documents)
        print("Indexed SharePoint pages:", result)
    except Exception as e:
        print(f"Error indexing SharePoint pages: {str(e)}")
def get_permissions(SHAREPOINT_DRIVE_ID,file_id, access_token):
    permissions_url = f"https://graph.microsoft.com/v1.0/drives/{SHAREPOINT_DRIVE_ID}/items/{file_id}/permissions"
    headers = {
        "Authorization": "Bearer " + access_token,
        "Accept": "application/json"
    }
    response = requests.get(permissions_url, headers=headers)
    permissions = []
    if response.status_code == 200:
        permissions_data = response.json().get("value", [])
        for permission_data in permissions_data:
            granted_to = permission_data.get("grantedTo", None)
            if granted_to:
                if "user" in granted_to:
                    user_or_group = granted_to["user"].get("displayName", "N/A")
                elif "group" in granted_to:
                    user_or_group = granted_to["group"].get("displayName", "N/A")
                if user_or_group != "N/A":  # Filter out "N/A" values
                    permissions.append(f" {user_or_group}")

    return permissions
def get_item_permissions(item_id,access_token,sharepoint_site_id,list_id):
    try:
        permissions_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/items/{item_id}/driveItem/permissions"
        headers = {"Authorization": "Bearer " + access_token, "Accept": "application/json"}
        response = requests.get(permissions_url, headers=headers)
        permissions = []
        
        if response.status_code == 200:
            permissions_data = response.json().get("value", [])
            for permission_data in permissions_data:
                granted_to = permission_data.get("grantedTo", None)
                if granted_to:
                    if "user" in granted_to:
                        user_or_group = granted_to["user"].get("displayName", "N/A")
                        permissions.append(user_or_group)
                    elif "group" in granted_to:
                        user_or_group = granted_to["group"].get("displayName", "N/A")
                        permissions.append(user_or_group)
        return permissions
    except Exception as e:
        print(f"Error fetching permissions for item {item_id}: {str(e)}")
        return []



def get_group_members(group_id, access_token):
    url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/transitiveMembers/"
    group_members = []
    headers = {
        "Authorization": "Bearer " + access_token,
        "Accept": "application/json"
    }

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        members = response.json()["value"]

        for member in members:
            principal_name = member["userPrincipalName"]
            group_members.append(principal_name)

    else:
        print(f"Failed to retrieve the group members: {response.text}")

    return group_members


def extract_text_from_pdf(file_content):
    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def extract_text_from_word(file_content):
    doc = Document(io.BytesIO(file_content))
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text


def extract_text_from_excel(file_content):
    file_str = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        file_str += str(cell.value) + " "
    except Exception as e:
        file_str = f"Error extracting text from Excel file: {str(e)}"
    return file_str

def extract_text_from_text(file_content):
    try:
        return file_content.decode('utf-8')
    except UnicodeDecodeError:
        return str(file_content, errors='ignore')


def extract_text_from_other(file_content):
    # Placeholder for handling other file types
    return "Unsupported file type"


def extract_text_from_csv(file_content):
    file_str = ""
    try:
        # Decode the bytes content assuming utf-8 encoding
        content_str = file_content.decode('utf-8')
        # Parse CSV content
        reader = csv.reader(content_str.splitlines())
        for row in reader:
            for cell in row:
                file_str += cell + " "
    except Exception as e:
        file_str = f"Error extracting text from CSV file: {str(e)}"
    return file_str

def index_folder_to_azure_search(folder_name, folder_id):
    client = SearchClient(endpoint=search_service_endpoint,
                          index_name=index_name,
                          credential=AzureKeyCredential(admin_key))

    document = {
        "id": str(folder_id),  # Unique identifier for the folder
        "folderName": folder_name
        # Add more fields as needed
    }

    try:
        result = client.upload_documents(documents=[document])
        print(f"Folder '{folder_name}' indexed successfully.")
    except Exception as e:
        print(
            f"Failed to index folder '{folder_name}' to Azure Search. Error: {str(e)}")
def index_document_to_azure_search(file_name, file_content, file_id, permissions, group_members, web_url):
    file_extension = os.path.splitext(file_name)[1].lower()
    extractors = {".docx": extract_text_from_word, ".pdf": extract_text_from_pdf, ".txt": extract_text_from_text, ".xlsx": extract_text_from_excel, ".xls": extract_text_from_excel, ".csv": extract_text_from_csv}
    content_str = extractors.get(file_extension, extract_text_from_other)(file_content)
    
    try:
        openai_credential = DefaultAzureCredential()
        token_provider = get_bearer_token_provider(openai_credential, "https://cognitiveservices.azure.com/.default")
        client = AzureOpenAI(azure_deployment=azure_openai_embedding_deployment, azure_endpoint=azure_openai_endpoint, api_key=azure_openai_key, azure_ad_token_provider=token_provider if not azure_openai_key else None, api_version=azure_openai_version)
        content_response = client.embeddings.create(input=content_str, model=embedding_model_name)
        combined_embeddings = content_response.data[0].embedding

        search_client = SearchClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(admin_key))
        document = {"id": str(file_id), "fileName": file_name, "content": content_str, "permissions": permissions, "accessList": [str(member) for member in group_members], "webUrl": web_url, "contentEmbeddings": combined_embeddings}
        search_client.upload_documents(documents=[document])
        print(f"Document '{file_name}' indexed successfully.")
    
    except Exception as e:
        print(f"Error indexing document '{file_name}': {str(e)}")




def index_drive_folder(SHAREPOINT_DRIVE_ID, access_token, folder_id=None):
    drive_items = get_sharepoint_items(access_token, folder_id, SHAREPOINT_DRIVE_ID)
    for item in drive_items:
        item_name = item['name']
        item_id = item['id']
        web_url = item['webUrl']
        if 'folder' in item:
            # If item is a folder, recursively index its contents
            index_folder_to_azure_search(item_name, item_id)
            index_drive_folder(SHAREPOINT_DRIVE_ID, access_token, item_id)
        else:
            # If item is a file, index its content
            file_content = download_sharepoint_file(item_id, SHAREPOINT_DRIVE_ID, access_token)
            permissions = get_permissions(SHAREPOINT_DRIVE_ID, item_id, access_token)
            group_members = get_group_members(group_id, access_token)
            if file_content:
                # Extract text based on file type
                if item_name.endswith('.pdf'):
                    content_str = extract_text_from_pdf(file_content)
                elif item_name.endswith('.docx'):
                    content_str = extract_text_from_word(file_content)
                elif item_name.endswith('.xlsx') or item_name.endswith('.xls'):
                    content_str = extract_text_from_excel(file_content)
                else:
                    content_str = extract_text_from_other(file_content)
                
                # Generate content embeddings using Azure OpenAI
                index_document_to_azure_search(item_name, file_content, item_id, permissions, group_members, web_url)
def get_access_token(client_id, client_secret, tenant_id):
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }

    response = requests.post(token_url, data=data)
    if response.status_code == 200:
        return response.json().get("access_token")
    else:
        print(f"Failed to retrieve access token: {response.text}")
        return None
def upload_documents(search_service_endpoint, admin_key, index_name, documents):
    client = SearchClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(admin_key))
    
    # Convert numerical values to strings in document data
    for document in documents:
        for key, value in document.items():
            if isinstance(value, (int, float)):
                document[key] = str(value)
    
    client.upload_documents(documents)



#sharepoint list
def get_sharepoint_list_data(list_id, access_token, sharepoint_site_id):
    try:
        schema_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/columns"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(schema_url, headers=headers)
        if response.status_code == 200:
            schema_data = response.json()
            fields_to_expand_and_index = [column["name"] for column in schema_data["value"]]


            select_fields = ','.join(fields_to_expand_and_index)
            url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists/{list_id}/items?expand=fields(select={select_fields})"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                if 'value' in data:
                    return data['value'], fields_to_expand_and_index
                else:
                    print("No items found in the SharePoint list.")
                    return [], []
            else:
                print(f"Failed to retrieve data from SharePoint list. Status code: {response.status_code}")
                return [], []
        else:
            print(f"Failed to retrieve schema from SharePoint list. Status code: {response.status_code}")
            return [], []
    except Exception as e:
        print(f"Error fetching data from SharePoint list: {str(e)}")
        return [], []
def get_all_sharepoint_lists(access_token, sharepoint_site_id):
    try:
        lists_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site_id}/lists"
        headers = {
            "Authorization": "Bearer " + access_token,
            "Accept": "application/json"
        }
        response = requests.get(lists_url, headers=headers)
        if response.status_code == 200:
            lists_data = response.json().get("value", [])
            all_lists_data = []
            for list_item in lists_data:
                list_id = list_item.get("id")
                list_data, fields = get_sharepoint_list_data(list_id, access_token, sharepoint_site_id)
                all_lists_data.append({"list_id": list_id, "list_data": list_data, "fields": fields})
            return all_lists_data
        else:
            print(f"Failed to retrieve lists from SharePoint site. Status code: {response.status_code}")
            return []
    except Exception as e:
        print(f"Error fetching lists from SharePoint site: {str(e)}")
        return []

def index_documents_to_azure_search(all_documents,group_id, item_id, access_token, sharepoint_site_id, list_id):
    try:
        openai_credential = DefaultAzureCredential()
        token_provider = get_bearer_token_provider(openai_credential, "https://cognitiveservices.azure.com/.default")
        client = AzureOpenAI(
            azure_deployment=azure_openai_embedding_deployment,
            azure_endpoint=azure_openai_endpoint,
            api_key=azure_openai_key,
            azure_ad_token_provider=token_provider if not azure_openai_key else None,
            api_version=azure_openai_version
        )
        client_search = SearchClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(admin_key))
        client_index = SearchIndexClient(endpoint=search_service_endpoint, index_name=index_name, credential=AzureKeyCredential(admin_key))
        
         # Fetch the index schema
        index_fields = [field.name for field in client_index.get_index(index_name).fields]
        access_list = get_group_members(group_id,access_token)
        # print("access_list",access_list)
        permissions = get_item_permissions(item_id, access_token, sharepoint_site_id, list_id)
        # print("permissions",permissions)
        for document in all_documents:
            document = {k: v for k, v in document.items() if k in index_fields}
            content = document.get('Content', '')
            document['accessList']= access_list
            document['Permissions'] = permissions
            embeddings = []
            chunk_size=8000
            content_str = "\n".join(content.split('\n'))
            content_chunks = [content_str[i:i + chunk_size] for i in range(0, len(content_str), chunk_size)]
            for chunk in content_chunks:
                try:
                    response = client.embeddings.create(input=chunk, model="text-embedding-ada-002")
                    embeddings = response.data[0].embedding[:1536]
                    embeddings_str = [str(embedding) for embedding in embeddings]
                    document['ContentEmbeddings'] = embeddings_str
                except Exception as e:
                    print(f"Error generating embeddings for content '{content_str}': {e}")
                print()
                
            try:
                    client_search.upload_documents(documents=[document])
                    print(f"Document '{document['id']}' indexed successfully.")
            except Exception as e:
                    print(f"Failed to index document '{document['id']}': {str(e)}")
    except Exception as e:
        print(f"Failed to index documents to Azure Cognitive Search. Error: {str(e)}")


@app.post("/sharepoint/get-details/")
def get_sharepoint_details(item: SharePointItem):
    if not item.ACCESS_TOKEN:
        if not (item.CLIENT_ID and item.CLIENT_SECRET and item.TENANT_ID):
            raise HTTPException(status_code=400, detail="Access token or client credentials must be provided.")
        item.ACCESS_TOKEN = get_access_token(item.CLIENT_ID, item.CLIENT_SECRET, item.TENANT_ID)
   
    if item.configuration == "drive":
                
        if item.SHAREPOINT_DRIVE_ID:
            if item.folder_id:
                drive_items = get_sharepoint_items(item.ACCESS_TOKEN, item.folder_id, item.SHAREPOINT_DRIVE_ID)
            else:
                drive_items = get_sharepoint_items(item.ACCESS_TOKEN, None, item.SHAREPOINT_DRIVE_ID)
            
            # Extract fields dynamically from the files in the drive
            fields = ["content", "weburl", "fileName","folderName"]
            create_or_update_index(search_service_endpoint, admin_key, index_name, fields)
            index_drive_folder(item.SHAREPOINT_DRIVE_ID, item.ACCESS_TOKEN)
            
            return {"drive_items": drive_items}
        else:
            raise HTTPException(status_code=400, detail="Drive ID is required for 'drive' configuration.")
   
    elif item.configuration == "list":
        if item.list_id and item.sharepoint_site_id:
            # Index data for the specific list
            list_data, fields = get_sharepoint_list_data(item.list_id, item.ACCESS_TOKEN, item.sharepoint_site_id)
            if not list_data:
                raise HTTPException(status_code=404, detail="No items found in the SharePoint list.")
            
            # Create or update index with the obtained fields
            
            
            # Index list data into Azure Cognitive Search
            all_documents = []
            for list_item in list_data:
                document = {"id": str(list_item["id"])}
                print("id",list_item["id"])
               
                fields_data = list_item.get("fields", {})
                if not isinstance(fields_data, dict):
                        print(f"Unexpected fields data structure for list_item {list_item['id']}: {fields_data}")
                        continue
                for field in fields:
                        rexgufield = regularexpression(field)
                        document[rexgufield] = str(list_item.get('fields', {}).get(field, ''))
                        print("Document fields:",document[rexgufield])
                all_documents.append(document)
                # print("all_documents",all_documents)
            create_or_update_index(search_service_endpoint, admin_key, index_name, fields)
            index_documents_to_azure_search(all_documents, group_id, list_item["id"], item.ACCESS_TOKEN, item.sharepoint_site_id, item.list_id)
            # print("all_documents",index_documents_to_azure_search(all_documents))

            return {"message": f"List '{item.list_id}' indexed successfully.", }
        elif item.sharepoint_site_id:
         # Index data for all lists within the SharePoint site
            all_lists_data = get_all_sharepoint_lists(item.ACCESS_TOKEN, item.sharepoint_site_id)
    
            if not all_lists_data:
                raise HTTPException(status_code=404, detail="No lists found in the SharePoint site.")

            # Iterate over all lists, retrieve their data and fields, and index them
            all_documents = []
            for list_data_info in all_lists_data:
                list_id = list_data_info["list_id"]
                print("Processing list_id:", list_id)
                list_data,fields= get_sharepoint_list_data(list_id, item.ACCESS_TOKEN, item.sharepoint_site_id)
                # print("fields",fields)
                # fields = list_data_info["fields"]
                sanitized_fields = sanitize_field_names(fields)
                

                # Retrieve the list data and fields
                               # print("List data:", list_data)
                # print("Fields:", fields)
                
                if not fields:
                    print(f"No fields found for list_id: {list_id}")
                    continue

                # Index list data into Azure Cognitive Search
                for list_item in list_data:
                    
                    document = {"id": (f"{list_id}_{list_item['id']}")}
                    # print("Processing list item:", list_item)

                    fields_data = list_item.get('fields', {})
                    if not isinstance(fields_data, dict):
                        print(f"Unexpected fields data structure for list_item {list_item['id']}: {fields_data}")
                        continue
                    for field in fields:
                        rexgufield = regularexpression(field)
                        document[rexgufield] = str(fields_data.get(field, ''))
                            # print(f"Field: {field}, Field Value: {field_value}")
                        # print(f"Field: {field}:",document[rexgufield])

                    all_documents.append(document)

            # Index the documents into Azure Cognitive Search
           

            
            
            # create_or_update_index(search_service_endpoint, admin_key, index_name, fields)
            if all_documents:
                 create_or_update_index(search_service_endpoint, admin_key, index_name, sanitized_fields)
                 index_documents_to_azure_search(all_documents, group_id, list_item["id"], item.ACCESS_TOKEN, item.sharepoint_site_id, item.list_id)
            print("All documents indexed successfully.")
            return {"all_lists_data": all_documents}

        elif item.configuration == "pages":
            if item.sharepoint_site_id:
                pages = get_sharepoint_pages(item.sharepoint_site_id, item.ACCESS_TOKEN)
                print("pages de", pages)
                all_documents = []
                create_or_update_index(search_service_endpoint, admin_key, index_name, fields)
                for page in pages:
                    document = {"id": str(page["id"])}
                    for field in fields:
                        field_name = sanitize_field_name(field)
                        document[field_name] = str(page.get('fields', {}).get(field, ''))
                    all_documents.append(document)
                index_sharepoint_pages(search_service_endpoint, index_name, admin_key, all_documents)
                return {"file_contents": all_documents}
        else:
            raise HTTPException(status_code=400, detail="SharePoint site ID is required for 'pages' configuration.")

    else:
        raise HTTPException(status_code=400, detail="Invalid configuration type. Must be 'list', 'drive', or 'pages'.")
    

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
